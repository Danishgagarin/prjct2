from flask import Flask, render_template_string, request, send_file, jsonify
import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os

app = Flask(__name__)

# Constants
FAIL_GRADES = {"F", "FE", "Absent", "Withheld"}
PASS_GRADES = {"S", "A+", "A", "B+", "B", "C+", "C", "D", "P"}
ALL_GRADES = ["S", "A+", "A", "B+", "B", "C+", "C", "D", "P", "F"]

BRANCH_MAP = {
    "CS": "Computer Science & Engineering",
    "EC": "Electronics & Communication Engineering",
    "ME": "Mechanical Engineering",
    "CE": "Civil Engineering",
    "EE": "Electrical & Electronics Engineering",
    "SF": "Safety & Fire Engineering",
    "CC": "Computer Science (Cyber Security)",
    "AD": "Artificial Intelligence & Data Science",
    "CSBS": "Computer Science & Business Systems",
}

# Helper functions
def extract_admission_year(register_no):
    match = re.search(r"\d{2}", register_no)
    return match.group() if match else None

def extract_branch_code(register_no):
    # Support 2, 3, or 4-letter branch codes (e.g. CS, ECE, CSBS)
    match = re.search(r"[A-Z]+\d{2}([A-Z]{2,4})\d+", register_no)
    return match.group(1) if match else "UNKNOWN"

def get_branch_name(branch_code):
    return BRANCH_MAP.get(branch_code, branch_code)

def parse_subjects(subject_text):
    """
    Extract all SUBJECT(GRADE) pairs from a subject string.
    Handles multi-line entries that have been joined (pdfplumber may
    concatenate continuation lines with a space or newline).
    """
    subjects = []
    # Find every CODE(GRADE) pattern anywhere in the text
    for match in re.finditer(r"([A-Z]{2,4}[0-9]{3})\(([^)]+)\)", subject_text):
        subjects.append((match.group(1), match.group(2)))
    return subjects

def parse_student_data_csv(csv_content):
    student_details = {}
    lines = csv_content.strip().split('\n')
    for line in lines[1:]:  # Skip header
        parts = [p.strip() for p in line.split(',')]
        if len(parts) >= 4:
            reg_no = parts[0]
            student_details[reg_no] = {
                'name': parts[1],
                'admission_category': parts[2],
                'scholar_type': parts[3]
            }
    return student_details

def apply_excel_formatting(file_path):
    wb = load_workbook(file_path)

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format headers (row 1)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Check for a second header row (used in Faculty_Branch_Subject sheet)
        if ws.max_row > 1:
            second_row_vals = [c.value for c in ws[2] if c.value]
            if second_row_vals and isinstance(second_row_vals[0], str) and second_row_vals[0].startswith("—"):
                for cell in ws[2]:
                    cell.fill = subheader_fill
                    cell.font = subheader_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Apply borders and grade coloring
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                if cell.value in PASS_GRADES:
                    cell.fill = pass_fill
                    cell.font = Font(bold=True, color="006100")
                elif cell.value in FAIL_GRADES:
                    cell.fill = fail_fill
                    cell.font = Font(bold=True, color="9C0006")

        ws.freeze_panes = "A2"

    wb.save(file_path)


def process_pdf(pdf_file, admission_year, faculty_mapping_text, student_data_csv):
    students = []
    faculty_map = {}
    student_details = {}

    # Parse faculty mapping
    if faculty_mapping_text:
        for line in faculty_mapping_text.strip().split('\n'):
            if '=' in line:
                subject, faculty = line.split('=', 1)
                faculty_map[subject.strip()] = faculty.strip()

    # Parse student details CSV
    if student_data_csv:
        student_details = parse_student_data_csv(student_data_csv)

    # Extract student data from PDF
    REGISTER_RE = re.compile(r"^[A-Z]+[0-9]{2}[A-Z]+[0-9]+")
    # A continuation line starts with a subject code pattern, NOT a register number
    CONTINUATION_RE = re.compile(r"^[A-Z]{2,4}[0-9]{3}\(")

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            # First pass: merge continuation lines onto their register line.
            # KTU PDFs wrap long subject lists onto the next line(s).
            raw_lines = text.split("\n")
            merged_lines = []
            for raw in raw_lines:
                raw = raw.strip()
                if not raw:
                    continue
                if CONTINUATION_RE.match(raw) and merged_lines:
                    # Append to the previous line
                    merged_lines[-1] = merged_lines[-1] + ", " + raw
                else:
                    merged_lines.append(raw)

            for line in merged_lines:
                if not REGISTER_RE.match(line):
                    continue

                parts = line.split(maxsplit=1)
                register_no = parts[0]

                if extract_admission_year(register_no) != admission_year:
                    continue

                branch_code = extract_branch_code(register_no)
                branch_name = get_branch_name(branch_code)

                subject_part = parts[1] if len(parts) > 1 else ""
                subjects = parse_subjects(subject_part)

                failed = sum(1 for _, g in subjects if g in FAIL_GRADES)
                overall = "PASS" if failed == 0 else "FAIL"

                details = student_details.get(register_no, {
                    'name': 'N/A',
                    'admission_category': 'N/A',
                    'scholar_type': 'N/A'
                })

                students.append({
                    "Branch": branch_name,
                    "Branch Code": branch_code,
                    "Register Number": register_no,
                    "Student Name": details['name'],
                    "Admission Category": details['admission_category'],
                    "Day Scholar/Hostler": details['scholar_type'],
                    "Failed Papers": failed,
                    "Overall Result": overall,
                    "_subjects_dict": {code: grade for code, grade in subjects}
                })

    students.sort(key=lambda x: (x["Branch Code"], x["Register Number"]))

    # Get all unique subjects
    all_subjects = sorted(set(
        subj for s in students for subj in s["_subjects_dict"].keys()
    ))

    # ─────────────────────────────────────────────
    # 1. Register-style sheets per branch
    # ─────────────────────────────────────────────
    branch_dataframes = {}
    for student in students:
        branch = student["Branch"]
        # Only include subjects that appear for this branch
        branch_subjects = sorted(set(
            subj for s in students if s["Branch"] == branch
            for subj in s["_subjects_dict"].keys()
        ))

        if branch not in branch_dataframes:
            branch_dataframes[branch] = {"rows": [], "subjects": branch_subjects}

        row_data = {
            "Sl No": len(branch_dataframes[branch]["rows"]) + 1,
            "Register Number": student["Register Number"],
            "Student Name": student["Student Name"],
            "Admission Category": student["Admission Category"],
            "Day Scholar/Hostler": student["Day Scholar/Hostler"]
        }
        for subject in branch_subjects:
            row_data[subject] = student["_subjects_dict"].get(subject, "-")
        row_data["Failed Papers"] = student["Failed Papers"]
        row_data["Overall Result"] = student["Overall Result"]

        branch_dataframes[branch]["rows"].append(row_data)

    branch_dfs = {
        branch: pd.DataFrame(info["rows"])
        for branch, info in branch_dataframes.items()
    }

    # ─────────────────────────────────────────────
    # 2. Overall subject analysis (aggregate)
    # ─────────────────────────────────────────────
    subject_analysis = {}
    for s in students:
        for subject, grade in s["_subjects_dict"].items():
            if subject not in subject_analysis:
                subject_analysis[subject] = {"Appeared": 0, "Passed": 0, "Failed": 0,
                                             **{g: 0 for g in ALL_GRADES}}
            subject_analysis[subject]["Appeared"] += 1
            if grade in subject_analysis[subject]:
                subject_analysis[subject][grade] += 1
            if grade in FAIL_GRADES:
                subject_analysis[subject]["Failed"] += 1
            else:
                subject_analysis[subject]["Passed"] += 1

    df_subjects_overall = pd.DataFrame([
        {
            "Subject Code": subject,
            "Students Appeared": data["Appeared"],
            "Students Passed": data["Passed"],
            "Students Failed": data["Failed"],
            "Pass Percentage": round((data["Passed"] / data["Appeared"] * 100), 2) if data["Appeared"] else 0,
            **{g: data[g] for g in ALL_GRADES}
        }
        for subject, data in subject_analysis.items()
    ])

    # ─────────────────────────────────────────────
    # 3. NEW: Branch-wise subject analysis
    #    For each subject, show results broken down by branch
    # ─────────────────────────────────────────────
    # Structure: {subject_code: {branch_name: {Appeared, Passed, Failed, grades...}}}
    subject_branch_analysis = {}
    for s in students:
        branch = s["Branch"]
        for subject, grade in s["_subjects_dict"].items():
            if subject not in subject_branch_analysis:
                subject_branch_analysis[subject] = {}
            if branch not in subject_branch_analysis[subject]:
                subject_branch_analysis[subject][branch] = {
                    "Appeared": 0, "Passed": 0, "Failed": 0,
                    **{g: 0 for g in ALL_GRADES}
                }
            subject_branch_analysis[subject][branch]["Appeared"] += 1
            if grade in subject_branch_analysis[subject][branch]:
                subject_branch_analysis[subject][branch][grade] += 1
            if grade in FAIL_GRADES:
                subject_branch_analysis[subject][branch]["Failed"] += 1
            else:
                subject_branch_analysis[subject][branch]["Passed"] += 1

    # Build rows for the branch-wise subject sheet
    branch_subject_rows = []
    for subject in sorted(subject_branch_analysis.keys()):
        branches_for_subject = subject_branch_analysis[subject]
        # Check if this subject spans multiple branches
        is_shared = len(branches_for_subject) > 1
        for branch, data in sorted(branches_for_subject.items()):
            branch_row = {
                "Subject Code": subject,
                "Branch": branch,
                "Students Appeared": data["Appeared"],
                "Students Passed": data["Passed"],
                "Students Failed": data["Failed"],
                "Pass Percentage": round((data["Passed"] / data["Appeared"] * 100), 2) if data["Appeared"] else 0,
                "Shared Subject": "Yes" if is_shared else "No",
            }
            for g in ALL_GRADES:
                branch_row[g] = data[g]
            branch_subject_rows.append(branch_row)

    df_subject_branch = pd.DataFrame(branch_subject_rows) if branch_subject_rows else pd.DataFrame()

    # ─────────────────────────────────────────────
    # 4. Faculty analysis — now with per-branch breakdown
    # ─────────────────────────────────────────────
    faculty_rows = []
    for subject in sorted(subject_analysis.keys()):
        faculty_name = faculty_map.get(subject, "Not Assigned")
        overall = subject_analysis[subject]
        # Overall row
        faculty_rows.append({
            "Subject Code": subject,
            "Faculty Name": faculty_name,
            "Branch": "ALL BRANCHES",
            "Students Appeared": overall["Appeared"],
            "Students Passed": overall["Passed"],
            "Students Failed": overall["Failed"],
            "Pass Percentage": round((overall["Passed"] / overall["Appeared"] * 100), 2) if overall["Appeared"] else 0,
        })
        # Per-branch rows (indented label)
        if subject in subject_branch_analysis:
            for branch, data in sorted(subject_branch_analysis[subject].items()):
                faculty_rows.append({
                    "Subject Code": f"  └─ {subject}",
                    "Faculty Name": faculty_name,
                    "Branch": branch,
                    "Students Appeared": data["Appeared"],
                    "Students Passed": data["Passed"],
                    "Students Failed": data["Failed"],
                    "Pass Percentage": round((data["Passed"] / data["Appeared"] * 100), 2) if data["Appeared"] else 0,
                })

    df_faculty = pd.DataFrame(faculty_rows) if faculty_rows else pd.DataFrame()

    # ─────────────────────────────────────────────
    # 5. Branch summary
    # ─────────────────────────────────────────────
    branch_summary = {}
    for s in students:
        branch = s["Branch"]
        if branch not in branch_summary:
            branch_summary[branch] = {"Total": 0, "Pass": 0, "Fail": 0}
        branch_summary[branch]["Total"] += 1
        if s["Overall Result"] == "PASS":
            branch_summary[branch]["Pass"] += 1
        else:
            branch_summary[branch]["Fail"] += 1

    df_branch_summary = pd.DataFrame([
        {
            "Branch": branch,
            "Total Students": data["Total"],
            "Passed": data["Pass"],
            "Failed": data["Fail"],
            "Pass Percentage": round((data["Pass"] / data["Total"] * 100), 2) if data["Total"] else 0
        }
        for branch, data in branch_summary.items()
    ])

    # ─────────────────────────────────────────────
    # Write to Excel
    # ─────────────────────────────────────────────
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1. Branch summary
        df_branch_summary.to_excel(writer, sheet_name="Branch_Summary", index=False)

        # 2. Register-style per branch
        for branch, df in branch_dfs.items():
            safe_name = branch.replace(" & ", "_").replace(" ", "_")[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

        # 3. Overall subject analysis
        df_subjects_overall.to_excel(writer, sheet_name="Subject_Analysis_Overall", index=False)

        # 4. NEW — Subject breakdown by branch
        if not df_subject_branch.empty:
            df_subject_branch.to_excel(writer, sheet_name="Subject_By_Branch", index=False)

        # 5. Faculty analysis (with branch breakdown)
        if not df_faculty.empty:
            df_faculty.to_excel(writer, sheet_name="Faculty_Analysis", index=False)

    output.seek(0)

    # Save temporarily for formatting
    temp_path = "temp_output.xlsx"
    with open(temp_path, 'wb') as f:
        f.write(output.read())

    apply_excel_formatting(temp_path)

    with open(temp_path, 'rb') as f:
        formatted_output = io.BytesIO(f.read())

    os.remove(temp_path)
    formatted_output.seek(0)
    return formatted_output


# ─────────────────────────────────────────────────────────────────────────────
# HTML Template — clean, minimal, professional redesign
# ─────────────────────────────────────────────────────────────────────────────
HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>KTU Result Analyzer</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg: #F5F4F0;
            --surface: #FFFFFF;
            --border: #E0DEDA;
            --border-strong: #C8C4BE;
            --text-primary: #1A1916;
            --text-secondary: #6B6860;
            --text-muted: #9C9A96;
            --accent: #2563EB;
            --accent-light: #EFF4FF;
            --success: #16A34A;
            --success-light: #F0FDF4;
            --error: #DC2626;
            --error-light: #FEF2F2;
            --radius: 6px;
            --mono: 'IBM Plex Mono', monospace;
        }

        * { margin: 0; padding: 0; box-sizing: border-box; }

        body {
            font-family: 'IBM Plex Sans', sans-serif;
            background: var(--bg);
            color: var(--text-primary);
            min-height: 100vh;
            font-size: 14px;
            line-height: 1.6;
        }

        /* Layout */
        .layout {
            display: grid;
            grid-template-columns: 260px 1fr;
            min-height: 100vh;
        }

        /* Sidebar */
        .sidebar {
            background: var(--text-primary);
            color: white;
            padding: 32px 24px;
            display: flex;
            flex-direction: column;
            gap: 32px;
            position: sticky;
            top: 0;
            height: 100vh;
            overflow-y: auto;
        }

        .sidebar-brand {
            display: flex;
            flex-direction: column;
            gap: 4px;
        }

        .sidebar-brand .label {
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            color: rgba(255,255,255,0.4);
        }

        .sidebar-brand h1 {
            font-size: 18px;
            font-weight: 600;
            color: white;
            line-height: 1.3;
        }

        .sidebar-divider {
            height: 1px;
            background: rgba(255,255,255,0.1);
        }

        .sidebar-section-label {
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 0.1em;
            text-transform: uppercase;
            color: rgba(255,255,255,0.35);
            margin-bottom: 10px;
        }

        .output-item {
            display: flex;
            align-items: flex-start;
            gap: 10px;
            padding: 10px 0;
            border-bottom: 1px solid rgba(255,255,255,0.08);
        }

        .output-item:last-child { border-bottom: none; }

        .output-dot {
            width: 6px;
            height: 6px;
            background: var(--accent);
            border-radius: 50%;
            margin-top: 6px;
            flex-shrink: 0;
        }

        .output-item-text {
            font-size: 12px;
            color: rgba(255,255,255,0.65);
            line-height: 1.5;
        }

        .output-item-text strong {
            display: block;
            color: rgba(255,255,255,0.9);
            font-weight: 500;
            margin-bottom: 1px;
        }

        /* Main content */
        .main {
            padding: 48px 56px;
            max-width: 720px;
        }

        .page-title {
            margin-bottom: 40px;
        }

        .page-title h2 {
            font-size: 24px;
            font-weight: 600;
            color: var(--text-primary);
            margin-bottom: 4px;
        }

        .page-title p {
            color: var(--text-secondary);
            font-size: 14px;
        }

        /* Form */
        .form-section {
            margin-bottom: 32px;
        }

        .section-label {
            font-size: 11px;
            font-weight: 600;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 16px;
            padding-bottom: 8px;
            border-bottom: 1px solid var(--border);
        }

        .field {
            margin-bottom: 20px;
        }

        .field label {
            display: block;
            font-size: 13px;
            font-weight: 500;
            color: var(--text-primary);
            margin-bottom: 6px;
        }

        .field .hint {
            font-size: 12px;
            color: var(--text-muted);
            margin-top: 5px;
        }

        /* File drop zone */
        .drop-zone {
            border: 1.5px dashed var(--border-strong);
            border-radius: var(--radius);
            padding: 24px;
            text-align: center;
            cursor: pointer;
            position: relative;
            transition: border-color 0.2s, background 0.2s;
            background: var(--surface);
        }

        .drop-zone:hover, .drop-zone.active {
            border-color: var(--accent);
            background: var(--accent-light);
        }

        .drop-zone input[type="file"] {
            position: absolute;
            inset: 0;
            opacity: 0;
            cursor: pointer;
            width: 100%;
            height: 100%;
        }

        .drop-zone-icon {
            font-size: 22px;
            margin-bottom: 6px;
        }

        .drop-zone-text {
            font-size: 13px;
            color: var(--text-secondary);
        }

        .drop-zone-text strong {
            color: var(--accent);
            font-weight: 500;
        }

        .drop-zone-filename {
            font-family: var(--mono);
            font-size: 12px;
            color: var(--accent);
            margin-top: 6px;
            font-weight: 500;
        }

        /* Text input */
        .text-input {
            width: 100%;
            padding: 10px 12px;
            border: 1.5px solid var(--border);
            border-radius: var(--radius);
            font-family: var(--mono);
            font-size: 13px;
            color: var(--text-primary);
            background: var(--surface);
            transition: border-color 0.2s;
            outline: none;
        }

        .text-input:focus {
            border-color: var(--accent);
        }

        .text-input.short {
            max-width: 120px;
            font-family: 'IBM Plex Sans', sans-serif;
        }

        textarea.text-input {
            resize: vertical;
            min-height: 100px;
            line-height: 1.6;
        }

        /* Code block hint */
        .code-hint {
            background: #F8F7F5;
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 10px 12px;
            font-family: var(--mono);
            font-size: 11.5px;
            color: var(--text-secondary);
            margin-top: 8px;
            line-height: 1.7;
        }

        /* Submit button */
        .submit-btn {
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            width: 100%;
            padding: 12px 24px;
            background: var(--text-primary);
            color: white;
            border: none;
            border-radius: var(--radius);
            font-family: 'IBM Plex Sans', sans-serif;
            font-size: 14px;
            font-weight: 500;
            cursor: pointer;
            transition: background 0.2s, opacity 0.2s;
            margin-top: 8px;
        }

        .submit-btn:hover { background: #2D2C29; }

        .submit-btn:disabled {
            opacity: 0.45;
            cursor: not-allowed;
        }

        /* Status messages */
        .status {
            display: none;
            align-items: center;
            gap: 10px;
            padding: 14px 16px;
            border-radius: var(--radius);
            font-size: 13px;
            margin-top: 16px;
            font-weight: 500;
        }

        .status.success {
            background: var(--success-light);
            color: var(--success);
            border: 1px solid #BBF7D0;
        }

        .status.error {
            background: var(--error-light);
            color: var(--error);
            border: 1px solid #FECACA;
        }

        /* Spinner */
        .spinner {
            width: 16px;
            height: 16px;
            border: 2px solid rgba(255,255,255,0.3);
            border-top-color: white;
            border-radius: 50%;
            animation: spin 0.7s linear infinite;
            display: none;
        }

        @keyframes spin {
            to { transform: rotate(360deg); }
        }

        .btn-content { display: flex; align-items: center; gap: 8px; }

        /* Responsive */
        @media (max-width: 768px) {
            .layout { grid-template-columns: 1fr; }
            .sidebar { height: auto; position: relative; }
            .main { padding: 32px 24px; }
        }
    </style>
</head>
<body>
<div class="layout">

    <!-- Sidebar -->
    <aside class="sidebar">
        <div class="sidebar-brand">
            <span class="label">KTU · Academic Tools</span>
            <h1>Result Analyzer</h1>
        </div>

        <div class="sidebar-divider"></div>

        <div>
            <p class="sidebar-section-label">Excel Output</p>
            <div class="output-item">
                <div class="output-dot"></div>
                <div class="output-item-text">
                    <strong>Branch Summary</strong>
                    Pass/fail totals per branch
                </div>
            </div>
            <div class="output-item">
                <div class="output-dot"></div>
                <div class="output-item-text">
                    <strong>Branch Register Sheets</strong>
                    Per-student grades, one sheet per branch
                </div>
            </div>
            <div class="output-item">
                <div class="output-dot"></div>
                <div class="output-item-text">
                    <strong>Subject Analysis (Overall)</strong>
                    Aggregate pass/fail across all branches
                </div>
            </div>
            <div class="output-item">
                <div class="output-dot"></div>
                <div class="output-item-text">
                    <strong>Subject Breakdown by Branch</strong>
                    Per-subject results split branch-wise — CS vs ME vs ECE separately
                </div>
            </div>
            <div class="output-item">
                <div class="output-dot"></div>
                <div class="output-item-text">
                    <strong>Faculty Analysis</strong>
                    Subject performance with per-branch rows per faculty
                </div>
            </div>
        </div>

        <div style="margin-top: auto;">
            <p style="font-size: 11px; color: rgba(255,255,255,0.25); line-height: 1.6;">
                Supports EKC / LEKC register formats.<br>
                Grades: S · A+ · A · B+ · B · C+ · C · D · P · F
            </p>
        </div>
    </aside>

    <!-- Main -->
    <main class="main">
        <div class="page-title">
            <h2>Generate Result Report</h2>
            <p>Upload a KTU result PDF to produce a formatted Excel analysis.</p>
        </div>

        <form id="uploadForm" enctype="multipart/form-data">

            <div class="form-section">
                <p class="section-label">Required</p>

                <div class="field">
                    <label>Result PDF</label>
                    <div class="drop-zone" id="pdfZone">
                        <div class="drop-zone-icon">📄</div>
                        <div class="drop-zone-text"><strong>Click to upload</strong> or drag & drop</div>
                        <div class="drop-zone-text" style="font-size:11px; margin-top:3px;">PDF files only</div>
                        <div class="drop-zone-filename" id="pdfFilename"></div>
                        <input type="file" name="pdf_file" id="pdfFile" accept=".pdf" required>
                    </div>
                </div>

                <div class="field">
                    <label>Admission Year</label>
                    <input type="text" class="text-input short" name="admission_year"
                           id="admissionYear" value="22" required placeholder="22">
                    <p class="hint">Last two digits — e.g. <code>22</code> for EKC<strong>22</strong>CS001</p>
                </div>
            </div>

            <div class="form-section">
                <p class="section-label">Optional</p>

                <div class="field">
                    <label>Student Details CSV</label>
                    <div class="drop-zone" id="csvZone">
                        <div class="drop-zone-icon">📋</div>
                        <div class="drop-zone-text"><strong>Click to upload</strong> CSV file</div>
                        <div class="drop-zone-filename" id="csvFilename"></div>
                        <input type="file" name="student_csv" id="studentCsv" accept=".csv">
                    </div>
                    <div class="code-hint">
                        Register Number,Name,Admission Category,Day Scholar/Hostler<br>
                        EKC22CS001,John Doe,Merit,Day Scholar<br>
                        LEKC22CS030,Jane Smith,Management,Hostler
                    </div>
                </div>

                <div class="field">
                    <label>Faculty Mapping</label>
                    <textarea class="text-input" name="faculty_mapping" id="facultyMapping"
                              placeholder="CST302=Dr. Anitha Krishnan&#10;CST304=Prof. Rajesh Nair&#10;HUT300=Dr. Meera Pillai"></textarea>
                    <p class="hint">One mapping per line: <code>SUBJECT_CODE=Faculty Name</code></p>
                </div>
            </div>

            <button type="submit" class="submit-btn" id="submitBtn">
                <span class="btn-content">
                    <span class="spinner" id="spinner"></span>
                    <span id="btnLabel">Generate Excel Report</span>
                </span>
            </button>
        </form>

        <div class="status success" id="successMessage">
            ✓ Report generated — download starting automatically.
        </div>
        <div class="status error" id="errorMessage"></div>

    </main>
</div>

<script>
    // File input display
    function bindFileInput(inputId, filenameId, zoneId) {
        const input = document.getElementById(inputId);
        const label = document.getElementById(filenameId);
        const zone = document.getElementById(zoneId);

        input.addEventListener('change', function () {
            if (this.files.length > 0) {
                label.textContent = this.files[0].name;
                zone.classList.add('active');
            } else {
                label.textContent = '';
                zone.classList.remove('active');
            }
        });
    }

    bindFileInput('pdfFile', 'pdfFilename', 'pdfZone');
    bindFileInput('studentCsv', 'csvFilename', 'csvZone');

    // Form submit
    const form = document.getElementById('uploadForm');
    const submitBtn = document.getElementById('submitBtn');
    const spinner = document.getElementById('spinner');
    const btnLabel = document.getElementById('btnLabel');
    const successMessage = document.getElementById('successMessage');
    const errorMessage = document.getElementById('errorMessage');

    form.addEventListener('submit', async function (e) {
        e.preventDefault();

        submitBtn.disabled = true;
        spinner.style.display = 'block';
        btnLabel.textContent = 'Processing…';
        successMessage.style.display = 'none';
        errorMessage.style.display = 'none';

        const formData = new FormData();
        formData.append('pdf_file', document.getElementById('pdfFile').files[0]);
        formData.append('admission_year', document.getElementById('admissionYear').value);
        formData.append('faculty_mapping', document.getElementById('facultyMapping').value);

        const csvInput = document.getElementById('studentCsv');
        if (csvInput.files.length > 0) {
            const csvText = await csvInput.files[0].text();
            formData.append('student_csv', csvText);
        }

        try {
            const response = await fetch('/analyze', { method: 'POST', body: formData });

            if (response.ok) {
                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = 'KTU_Result_Analysis.xlsx';
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                document.body.removeChild(a);
                successMessage.style.display = 'flex';
            } else {
                const error = await response.text();
                errorMessage.textContent = 'Error: ' + error;
                errorMessage.style.display = 'flex';
            }
        } catch (err) {
            errorMessage.textContent = 'Network error. Please try again.';
            errorMessage.style.display = 'flex';
        } finally {
            spinner.style.display = 'none';
            btnLabel.textContent = 'Generate Excel Report';
            submitBtn.disabled = false;
        }
    });
</script>
</body>
</html>'''


@app.route('/')
def index():
    return HTML_TEMPLATE


@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        pdf_file = request.files['pdf_file']
        admission_year = request.form['admission_year']
        faculty_mapping = request.form.get('faculty_mapping', '')
        student_csv = request.form.get('student_csv', '')

        if not pdf_file:
            return "No PDF file uploaded", 400

        result_file = process_pdf(pdf_file, admission_year, faculty_mapping, student_csv)

        return send_file(
            result_file,
            as_attachment=True,
            download_name='KTU_Result_Analysis.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return str(e), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
